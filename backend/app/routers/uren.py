from io import BytesIO

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.database import async_session
from app.models.uren import TimeEntry, TimeSheet

router = APIRouter()


@router.get("/api/uren")
async def list_timesheets() -> list[dict]:
    """List all timesheets, newest first."""
    async with async_session() as session:
        stmt = (
            select(TimeSheet)
            .options(selectinload(TimeSheet.entries))
            .order_by(TimeSheet.datum.desc())
        )
        result = await session.execute(stmt)
        sheets = result.scalars().all()
        return [
            {
                "id": s.id,
                "datum": s.datum.isoformat(),
                "totaal_uren": _format_duur(sum(e.duur_minuten for e in s.entries)),
                "entries_count": len(s.entries),
                "created_at": s.created_at.isoformat(),
            }
            for s in sheets
        ]


@router.get("/api/uren/{sheet_id}")
async def get_timesheet(sheet_id: str) -> dict:
    """Get a single timesheet with all entries."""
    async with async_session() as session:
        stmt = (
            select(TimeSheet)
            .options(selectinload(TimeSheet.entries))
            .where(TimeSheet.id == sheet_id)
        )
        result = await session.execute(stmt)
        sheet = result.scalar_one_or_none()
        if not sheet:
            raise HTTPException(status_code=404, detail="Urenregistratie niet gevonden")
        total_min = sum(e.duur_minuten for e in sheet.entries)
        return {
            "id": sheet.id,
            "datum": sheet.datum.isoformat(),
            "totaal_uren": _format_duur(total_min),
            "entries_count": len(sheet.entries),
            "source_text": sheet.source_text,
            "created_at": sheet.created_at.isoformat(),
            "entries": [
                {
                    "start": e.start_tijd,
                    "eind": e.eind_tijd,
                    "omschrijving": e.omschrijving,
                    "duur_minuten": e.duur_minuten,
                }
                for e in sheet.entries
            ],
        }


@router.get("/api/uren/{sheet_id}/excel")
async def download_excel(sheet_id: str):
    """Generate and download an Excel file for a timesheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    async with async_session() as session:
        stmt = (
            select(TimeSheet)
            .options(selectinload(TimeSheet.entries))
            .where(TimeSheet.id == sheet_id)
        )
        result = await session.execute(stmt)
        sheet = result.scalar_one_or_none()
        if not sheet:
            raise HTTPException(status_code=404, detail="Niet gevonden")

    wb = Workbook()
    ws = wb.active
    ws.title = "Uren"

    headers = ["Start", "Eind", "Omschrijving", "Duur"]
    bold = Font(bold=True)
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = bold

    for i, entry in enumerate(sheet.entries, 2):
        ws.cell(row=i, column=1, value=entry.start_tijd)
        ws.cell(row=i, column=2, value=entry.eind_tijd)
        ws.cell(row=i, column=3, value=entry.omschrijving)
        ws.cell(row=i, column=4, value=_format_duur(entry.duur_minuten))

    total_row = len(sheet.entries) + 2
    total_min = sum(e.duur_minuten for e in sheet.entries)
    ws.cell(row=total_row, column=3, value="Totaal").font = bold
    ws.cell(row=total_row, column=4, value=_format_duur(total_min)).font = bold

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 8

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"uren_{sheet.datum.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _format_duur(minuten: int) -> str:
    return f"{minuten // 60}:{minuten % 60:02d}"
